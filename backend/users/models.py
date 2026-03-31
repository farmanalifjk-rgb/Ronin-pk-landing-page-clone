import uuid
from django.db import models
from django.contrib.auth.models import AbstractUser
from django.conf import settings
import os
from django.db.models.signals import post_save
from django.dispatch import receiver
from openpyxl import Workbook, load_workbook


class CustomUser(AbstractUser):

    # 2. Add this line to override the default integer ID with a UUID
    id = models.UUIDField(primary_key=True, default=uuid.uuid4, editable=False)


    email = models.EmailField(unique=True)

    # This makes Email the main login field
    USERNAME_FIELD = 'email'
    REQUIRED_FIELDS = ['username']

class Profile(models.Model):
    user = models.OneToOneField(settings.AUTH_USER_MODEL, on_delete=models.CASCADE)
    image = models.ImageField(default='default.jpg', upload_to='profile_pics')

    def __str__(self):
        return f'{self.user.username} Profile'
    

@receiver(post_save, sender=settings.AUTH_USER_MODEL)
def update_user_excel_file(sender, instance, **kwargs):
    # This will create 'users_data.xlsx' in your main project folder
    filepath = os.path.join(settings.BASE_DIR, 'users_data.xlsx')

    # 1. Load the Excel file (or create it if it doesn't exist)
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Create the header row for a brand new file
        ws.append(['ID', 'Username', 'Email'])

    # 2. Search for the user's ID in the Excel file
    # CRITICAL FIX: Convert the UUID to a string so Excel can read it!
    user_id = str(instance.id)

    row_to_update = None
    # We start at row 2 to skip the Headers
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == user_id:
            row_to_update = row
            break

    # 3. Update the existing row OR add a new row
    if row_to_update:
        ws.cell(row=row_to_update, column=2).value = instance.username
        ws.cell(row=row_to_update, column=3).value = instance.email
    else:
        # Save the STRING version of the UUID to Excel
        ws.append([user_id, instance.username, instance.email])
        print(f"Added {instance.username} to Excel!")


    # 4. Save the file (WITH CRASH PROTECTION)
    try:
        wb.save(filepath)
    except PermissionError:
        print(f"⚠️ WARNING: Could not update users_data.xlsx because the file is currently open in another program! Please close it.")